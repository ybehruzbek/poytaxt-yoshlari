"""Kanallar bazasi uchun Excel import/eksport (TZ 2.2).

Import ikki formatni tushunadi:

1) Sarlavhali fayl — birinchi 10 qatordan sarlavha qatori qidiriladi va
   ustunlar nomidan aniqlanadi: «Havola»/«Link», «OTT nomi», «Turi»,
   «Teg» (ixtiyoriy). Ustunlar tartibi va ortiqcha ustunlar (№, Kanal nomi
   va h.k.) ahamiyatsiz.

2) Sarlavhasiz oddiy format: A — OTT nomi, B — link, C — turi, D — teg.

Turi qiymatlari: Davlat / Xorijiy / Nodavlat (bo'sh bo'lsa Davlat).
"""
from io import BytesIO

from openpyxl import Workbook, load_workbook

from zakovat_bot.models import Channel, OttType

_TYPE_MAP = {
    "davlat": OttType.DAVLAT,
    "xorijiy": OttType.XORIJIY,
    "nodavlat": OttType.NODAVLAT,
}


def normalize_link(raw):
    """Havoladan @siz username yoki -100... chat_id ajratadi.
    Qaytaradi: (username, chat_id) — bittasi None bo'ladi; ikkalasi None = xato."""
    if raw is None:
        return None, None
    value = str(raw).strip()
    if not value:
        return None, None
    for prefix in ("https://t.me/", "http://t.me/", "t.me/"):
        if value.startswith(prefix):
            value = value[len(prefix):]
            break
    value = value.lstrip("@").strip().rstrip("/")
    if not value:
        return None, None
    if value.startswith("-100") and value[1:].isdigit():
        return None, int(value)
    if value.lstrip("-").isdigit():
        # boshqa raqamli ID ham qabul qilinadi
        return None, int(value)
    # username: faqat ruxsat etilgan belgilar (Telegram: 5-32 belgi)
    if all(c.isalnum() or c == "_" for c in value) and 5 <= len(value) <= 32:
        return value, None
    return None, None


def _cell(row, idx):
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


def _detect_columns(rows):
    """Birinchi 10 qatordan sarlavha qatorini qidiradi.

    Qaytaradi: (data_boshlanish_indeksi, {link, name, type, tag} ustun raqamlari).
    Sarlavha topilmasa — oddiy pozitsion format (A/B/C/D) deb olinadi.
    """
    for i, row in enumerate(rows[:10]):
        cols = {"link": None, "name": None, "type": None, "tag": None}
        for j, cell in enumerate(row or ()):
            text = str(cell).strip().lower() if cell is not None else ""
            if not text:
                continue
            if cols["link"] is None and ("havola" in text or "link" in text):
                cols["link"] = j
            elif cols["name"] is None and "ott" in text and "nomi" in text:
                cols["name"] = j
            elif cols["type"] is None and "tur" in text:
                cols["type"] = j
            elif cols["tag"] is None and ("teg" in text or "tag" in text):
                cols["tag"] = j
        # «OTT nomi» topilmasa, «kanal nomi» bo'lmagan oddiy «nomi» ham bo'ladi
        if cols["link"] is not None and cols["name"] is None:
            for j, cell in enumerate(row or ()):
                text = str(cell).strip().lower() if cell is not None else ""
                if "nomi" in text and "kanal" not in text:
                    cols["name"] = j
                    break
        if cols["link"] is not None and cols["name"] is not None:
            return i + 1, cols
    return 0, {"name": 0, "link": 1, "type": 2, "tag": 3}


def import_channels(file_bytes):
    """Excel'dan kanallarni bazaga qo'shadi/yangilaydi.

    Qaytaradi: dict(added, updated, duplicates, errors: list[(qator_no, sabab)]).
    Mavjud kanal ustiga yozilmaydi — faqat bo'sh maydonlari to'ldiriladi (TZ 2.2).
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = [row for row in ws.iter_rows(values_only=True)]
    wb.close()

    start, cols = _detect_columns(rows)

    added = updated = duplicates = 0
    errors = []
    seen_in_file = set()

    for row_no, row in enumerate(rows[start:], start=start + 1):
        if not row or all(cell is None or not str(cell).strip() for cell in row):
            continue

        ott_name = _cell(row, cols["name"])
        link_raw = _cell(row, cols["link"]) or None
        type_raw = _cell(row, cols["type"]).lower()
        tag = _cell(row, cols["tag"])

        username, chat_id = normalize_link(link_raw)
        if username is None and chat_id is None:
            errors.append((row_no, "link noto'g'ri yoki bo'sh"))
            continue
        if not ott_name:
            errors.append((row_no, "OTT nomi bo'sh"))
            continue

        key = username or chat_id
        if key in seen_in_file:
            duplicates += 1
            continue
        seen_in_file.add(key)

        ott_type = _TYPE_MAP.get(type_raw, OttType.DAVLAT)

        existing = None
        if chat_id:
            existing = Channel.all_objects.filter(chat_id=chat_id).first()
        if existing is None and username:
            existing = Channel.all_objects.filter(username__iexact=username).first()

        if existing:
            # Ustiga yozmaymiz — faqat bo'sh maydonlarni to'ldiramiz
            changed = []
            if not existing.username and username:
                existing.username = username
                changed.append("username")
            if not existing.chat_id and chat_id:
                existing.chat_id = chat_id
                changed.append("chat_id")
            if not existing.tag and tag:
                existing.tag = tag
                changed.append("tag")
            if changed:
                existing.save(update_fields=changed)
                updated += 1
            else:
                duplicates += 1
            continue

        Channel.objects.create(
            chat_id=chat_id,
            username=username,
            ott_name=ott_name,
            ott_type=ott_type,
            tag=tag or None,
        )
        added += 1

    return {"added": added, "updated": updated, "duplicates": duplicates, "errors": errors}


def export_channels():
    """Joriy bazani Excel-fayl (bytes) ko'rinishida qaytaradi (TZ 2.2)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Kanallar"
    ws.append([
        "OTT nomi", "Link", "Turi", "Teg",
        "Holati", "Bot admin", "Chat ID", "Qo'shilgan sana",
    ])

    for ch in Channel.all_objects.order_by("ott_name"):
        link = f"@{ch.username}" if ch.username else (str(ch.chat_id) if ch.chat_id else "")
        ws.append([
            ch.ott_name,
            link,
            ch.get_ott_type_display(),
            ch.tag or "",
            "faol" if ch.is_active else "nofaol",
            "ha" if ch.bot_is_admin else "yo'q",
            ch.chat_id or "",
            ch.created_datetime.strftime("%Y-%m-%d %H:%M"),
        ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_broadcast_report(broadcast):
    """Bitta tarqatish bo'yicha Excel hisobot (TZ 5)."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Tarqatish {broadcast.id}"
    ws.append(["OTT nomi", "Kanal", "Holati", "Xato sababi", "Yuborilgan vaqt"])

    for r in broadcast.results.select_related("channel").order_by("channel__ott_name"):
        ch = r.channel
        link = f"@{ch.username}" if ch.username else (str(ch.chat_id) if ch.chat_id else "")
        ws.append([
            ch.ott_name,
            link,
            "yetkazildi" if r.delivered else "xato",
            r.error_reason or "",
            r.sent_at.strftime("%Y-%m-%d %H:%M") if r.sent_at else "",
        ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
